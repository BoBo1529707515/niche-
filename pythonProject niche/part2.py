from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import time
import pandas as pd
from openpyxl import load_workbook
from selenium.common.exceptions import NoSuchElementException
# 大学名称
name_all = []
location_all = []
Overall_Niche_Grade_all = []
Academics_all = []
Diversity_all = []
Teachers_all = []
College_Prep_all = []
Clubs_Activities_all = []
Sports_all = []
com_all = []
tel_all = []
more_location_all = []
Application_Deadline_all = []
Application_Fee_all = []
Interview_Required_all = []
Required_Recommended_Tests_all = []
Best_Private_all = []
Best_College_all = []
Best_High_all = []
Average_Graduation_Rate_all = []
Average_SAT_all = []
Average_ACT_all = []
Students_all = []
Student_Teacher_Ratio_all = []




def get_text(xpath, target_list, default_value="无"):
    try:
        element = driver.find_element(By.XPATH, xpath)
        text = element.text
        target_list.append(text)
    except NoSuchElementException:
        target_list.append(default_value)
        text = default_value  # 给 text 变量赋予默认值
    print(text)



print("大学名称列表长度:", len(name_all))
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options
# 设置浏览器路径
edge_driver_path = r'F:\edgedriver_win32 (1)\msedgedriver.exe'



# 创建 EdgeOptions 对象并设置代理
edge_options = Options()


# 创建 Edge 服务对象
edge_driver_path = '/path/to/msedgedriver'
service = Service(r'F:\edgedriver_win32 (1)\msedgedriver.exe')

# 创建 Edge 浏览器对象
driver = webdriver.Edge(service=service, options=edge_options)

# 获取网页
file_path = r'C:\Users\15297\Desktop\niche初步2.xlsx'
wb = load_workbook(filename=file_path)
sheet = wb.active

# 爬取每个链接对应页面的信息
for row in range(2, 80):  # 读取前n行
    link = sheet.cell(row=row, column=3).value

    print("当前链接:", link)
    driver.get(link)
    get_text('//*[@id="header"]/div/div[2]/div[1]/h1', name_all)
    get_text('//*[@id="header"]/div/div[2]/div[1]/ul[1]/li[4]', location_all)
    get_text('//*[@id="report-card"]/div/div/div/div[1]/div/div/div[1]/div', Overall_Niche_Grade_all)
    get_text('//*[@id="report-card"]/div/div/div/div[2]/ol/li[1]/div/div[2]', Academics_all)
    get_text('//*[@id="report-card"]/div/div/div/div[2]/ol/li[2]/div/div[2]', Diversity_all)
    get_text('//*[@id="report-card"]/div/div/div/div[2]/ol/li[3]/div/div[2]', Teachers_all)
    get_text('//*[@id="report-card"]/div/div/div/div[2]/ol/li[4]/div/div[2]', College_Prep_all)
    get_text('//*[@id="report-card"]/div/div/div/div[2]/ol/li[5]/div/div[2]', Clubs_Activities_all)
    get_text('//*[@id="report-card"]/div/div/div/div[2]/ol/li[6]/div/div[2]', Sports_all)
    get_text('//*[@id="about"]/div[2]/div[1]/div/div[1]/div/a', com_all)
    get_text('//*[@id="about"]/div[2]/div[1]/div/div[2]/div/a', tel_all)
    get_text('//*[@id="about"]/div[2]/div[1]/div/div[3]/div/address', more_location_all)
    get_text('//*[@id="applying"]/div[2]/div[1]/div/div[1]/div[2]/span', Application_Deadline_all)
    get_text('//*[@id="applying"]/div[2]/div[1]/div/div[2]/div[2]/span', Application_Fee_all)
    get_text('//*[@id="applying"]/div[2]/div[1]/div/div[3]/div[2]/span', Interview_Required_all)
    get_text('//*[@id="applying"]/div[2]/div[1]/div/div[4]/div[2]/span', Required_Recommended_Tests_all)
    get_text('//*[@id="rankings"]/div/div[3]/div[1]/div/ul/li[1]/button/div[2]', Best_Private_all)
    get_text('//*[@id="rankings"]/div/div[3]/div[1]/div/ul/li[2]/button/div[2]', Best_College_all)
    get_text('//*[@id="rankings"]/div/div[3]/div[1]/div/ul/li[3]/button/div[2]', Best_High_all)
    get_text('//*[@id="academics"]/div[2]/div[1]/div/div[1]/div[2]/span', Average_Graduation_Rate_all)
    get_text('//*[@id="academics"]/div[2]/div[1]/div/div[2]/div[2]', Average_SAT_all)
    get_text('//*[@id="academics"]/div[2]/div[1]/div/div[3]/div[2]', Average_ACT_all)
    get_text('//*[@id="students"]/div[2]/div[1]/div/div[2]/div[2]/span', Students_all)
    get_text('//*[@id="teachers"]/div[2]/div[1]/div/div[1]/div[2]/span', Student_Teacher_Ratio_all)
    time.sleep(1)
data = pd.DataFrame({
    '名称':name_all,
    '位置':location_all,
    'Overall_Niche_Grade':Overall_Niche_Grade_all,
    'Academics':Academics_all,
    'Diversity':Diversity_all,
    'Teachers':Teachers_all,
    'College_Prep':College_Prep_all,
    'Clubs_Activities':Clubs_Activities_all,
    'Sports':Sports_all,
    'com':com_all,
    'tel':tel_all,
    'more_location':more_location_all,
    'Application_Deadline':Application_Deadline_all,
    'Application_Fee':Application_Fee_all,
    'Interview_Required_all':Interview_Required_all,
    'Required_Recommended_Tests':Required_Recommended_Tests_all,
    'Best Private High Schools in California':Best_Private_all,
    'Best College Prep Private High Schools in California':Best_College_all,
    'Best High Schools for STEM in California': Best_High_all,
    'Average_Graduation_Rate': Average_Graduation_Rate_all,
    'Average_SAT': Average_SAT_all,
    'Average_ACT': Average_ACT_all,
    'Students': Students_all,
    'Student_Teacher_Ratio': Student_Teacher_Ratio_all,
})
data.to_excel(r"C:\Users\15297\Desktop\niche结果texas.xlsx", index=False)





data.head()